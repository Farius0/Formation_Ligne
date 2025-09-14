# 📊 MANIPULATION DES FICHIERS EXCEL (.XLSX) AVEC openpyxl

import openpyxl  # Importation du module openpyxl pour manipuler les fichiers Excel

# 📌 Chargement des fichiers Excel avec `data_only=True` pour récupérer les valeurs calculées des formules
wb1 = openpyxl.load_workbook("octobre.xlsx", data_only=True)
wb2 = openpyxl.load_workbook("novembre.xlsx", data_only=True)
wb3 = openpyxl.load_workbook("decembre.xlsx", data_only=True)

# Dictionnaire pour stocker les données des ventes sous la forme :
# {"Pommes": [760, 660, 900], "Bananes": [500, 450, 600], ...}
donnees = {}

def ajouter_data_depuis_wb(wb, d):
    """
    Fonction pour extraire les données de ventes d'un fichier Excel et les ajouter dans un dictionnaire.

    :param wb: Classeur Excel ouvert (Workbook)
    :param d: Dictionnaire contenant les données agrégées
    """
    sheet = wb.active  # Sélection de la feuille active

    # 🔹 Parcours des lignes (en supposant que les données commencent à la ligne 2)
    for row in range(2, sheet.max_row + 1):  
        nom_article = sheet.cell(row, 1).value  # Lecture du nom de l'article (colonne A)
        total_ventes = sheet.cell(row, 4).value  # Lecture du total des ventes (colonne D)

        # Vérification que la cellule contenant le nom de l'article n'est pas vide
        if not nom_article:
            break  # Sort de la boucle si on atteint une ligne vide

        # 🔹 Ajout des ventes dans le dictionnaire
        if nom_article in d:
            d[nom_article].append(total_ventes)  # Ajout des ventes à la liste existante
        else:
            d[nom_article] = [total_ventes]  # Création d'une nouvelle entrée

# Ajout des données des trois fichiers dans le dictionnaire `donnees`
ajouter_data_depuis_wb(wb1, donnees)
ajouter_data_depuis_wb(wb2, donnees)
ajouter_data_depuis_wb(wb3, donnees)

# Affichage du dictionnaire contenant les ventes consolidées
print(donnees)
