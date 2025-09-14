# FICHIERS EXCEL - Lecture d'un fichier .XLSX avec openpyxl

import openpyxl  # Importation du module openpyxl pour manipuler les fichiers Excel (.xlsx)

# Chargement du fichier Excel "octobre.xlsx"
wb = openpyxl.load_workbook("octobre.xlsx")

# Affichage des noms des feuilles disponibles dans le fichier
print("Feuilles disponibles :", wb.sheetnames)

# Sélection de la première feuille du classeur (workbook)
sheet = wb[wb.sheetnames[0]]  # Ou bien : sheet = wb.active pour la feuille active

# 🔹 Lecture d'une cellule spécifique (exemple : cellule B7)
# cell = sheet["B7"]
# print("Valeur de la cellule B7 :", cell.value)

# 🔹 Obtenir les dimensions de la feuille
# print("Nombre total de lignes :", sheet.max_row)
# print("Nombre total de colonnes :", sheet.max_column)

# 🔹 Lecture des valeurs d'une colonne (exemple : de la ligne 2 à 6, colonne B)
# print("Lecture des valeurs de la colonne B (de la ligne 2 à 6) :")
# for row in range(2, 7):  # De la ligne 2 à 6
#     cell = sheet.cell(row, 2)  # Colonne 2 (B)
#     print(f"Ligne {row}, Colonne B :", cell.value)
