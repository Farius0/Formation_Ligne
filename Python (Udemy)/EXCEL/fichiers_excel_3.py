# 📊 MANIPULATION DES FICHIERS EXCEL (.XLSX) AVEC openpyxl

import openpyxl  # Importation du module openpyxl pour manipuler les fichiers Excel

# 📌 Chargement des fichiers Excel avec `data_only=True` pour récupérer les valeurs calculées des formules
wb1 = openpyxl.load_workbook("octobre.xlsx", data_only=True)
wb2 = openpyxl.load_workbook("novembre.xlsx", data_only=True)
wb3 = openpyxl.load_workbook("decembre.xlsx", data_only=True)

# Dictionnaire pour stocker les données des ventes sous la forme :
# {"Pommes": [760, 660, 900], "Bananes": [500, 450, 600], ...}
donnees = {}

def ajouter_data_depuis_wb(wb, d):
    """
    Fonction pour extraire les données de ventes d'un fichier Excel et les ajouter dans un dictionnaire.

    :param wb: Classeur Excel ouvert (Workbook)
    :param d: Dictionnaire contenant les données agrégées
    """
    sheet = wb.active  # Sélection de la feuille active

    # 🔹 Parcours des lignes (en supposant que les données commencent à la ligne 2)
    for row in range(2, sheet.max_row + 1):  
        nom_article = sheet.cell(row, 1).value  # Lecture du nom de l'article (colonne A)
        total_ventes = sheet.cell(row, 4).value  # Lecture du total des ventes (colonne D)

        # Vérification que la cellule contenant le nom de l'article n'est pas vide
        if not nom_article:
            break  # Sort de la boucle si on atteint une ligne vide

        # 🔹 Ajout des ventes dans le dictionnaire
        if nom_article in d:
            d[nom_article].append(total_ventes)  # Ajout des ventes à la liste existante
        else:
            d[nom_article] = [total_ventes]  # Création d'une nouvelle entrée

# Ajout des données des trois fichiers dans le dictionnaire `donnees`
ajouter_data_depuis_wb(wb1, donnees)
ajouter_data_depuis_wb(wb2, donnees)
ajouter_data_depuis_wb(wb3, donnees)

# Affichage du dictionnaire contenant les ventes consolidées
print(donnees)

# 📊 CRÉATION D'UN NOUVEAU FICHIER EXCEL POUR SAUVEGARDER LES DONNÉES AGRÉGÉES

# 📌 Création d'un nouveau classeur Excel (Workbook)
wb_sortie = openpyxl.Workbook()

# Sélection de la feuille active (par défaut, il y a une seule feuille)
sheet = wb_sortie.active

# 🔹 Ajout des en-têtes dans la première ligne du fichier Excel
sheet["A1"] = "Article"  # Colonne des articles
sheet["B1"] = "Octobre"  # Colonne des ventes en Octobre
sheet["C1"] = "Novembre"  # Colonne des ventes en Novembre
sheet["D1"] = "Décembre"  # Colonne des ventes en Décembre

# 🔹 Remplissage des données consolidées
row = 2  # Début des données à la ligne 2 (après les en-têtes)
for nom_article, ventes in donnees.items():  # Parcours du dictionnaire des ventes
    sheet.cell(row, 1).value = nom_article  # Ajout du nom de l'article en colonne A

    # Ajout des ventes mensuelles dans les colonnes correspondantes (B, C, D)
    for j in range(len(ventes)):
        sheet.cell(row, 2 + j).value = ventes[j]  # Placement des valeurs de ventes

    row += 1  # Passage à la ligne suivante

# 📌 Sauvegarde du fichier Excel consolidé
wb_sortie.save("total_ventes_trimestre.xlsx")

# ✅ Fichier Excel "total_ventes_trimestre.xlsx" correctement généré avec les données de ventes consolidées !
